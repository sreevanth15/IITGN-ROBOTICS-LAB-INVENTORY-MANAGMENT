#!/usr/bin/env python3
"""
Import inventory from an Excel file into the existing SQLite management DB.

Usage:
  python import_inventory_from_excel_py.py <file.xlsx> [--sheet SheetName] [--addedBy "Name"]

The script upserts products by `product_name` + `location`, creates/updates
`inventory_status`, and logs `PRODUCT_ADD` for new inserts in `transaction_history`.
"""
import os
import sqlite3
import argparse
from typing import List

try:
    from openpyxl import load_workbook
except Exception as e:
    raise SystemExit('Please install openpyxl (pip install openpyxl) to run this script')


DB_PATH = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', '..', 'database', 'inventory.db'))


def clean_key(s: str) -> str:
    if s is None:
        return ''
    return ''.join(ch for ch in str(s).lower() if ch.isalnum())


def build_key_map(headers: List[str]):
    return {clean_key(h): idx for idx, h in enumerate(headers)}


CANDIDATES = {
    'product_name': ['productname', 'product', 'item', 'name'],
    'description': ['description', 'desc', 'details'],
    'category': ['category', 'cat'],
    'unit': ['unit', 'uom'],
    'location': ['location', 'shelf', 'shelflocation', 'shelfno', 'shelfnumber'],
    'quantity_in_stock': ['quantity', 'qty', 'quantityinstock', 'instock', 'stock', 'count'],
    'quantity_in_use': ['quantityinuse', 'inuse', 'incirculation', 'circulation'],
    'quantity_total': ['quantitytotal', 'total', 'qtytotal']
}


def get_val_by_candidates(row: List, key_map: dict, candidates: List[str]):
    for cand in candidates:
        nk = ''.join(ch for ch in cand.lower() if ch.isalnum())
        if nk in key_map:
            val = row[key_map[nk]]
            return val
    return None


def to_int(v):
    if v is None or v == '':
        return None
    try:
        return int(v)
    except Exception:
        try:
            return int(float(v))
        except Exception:
            return None


def upsert_row(conn: sqlite3.Connection, row_vals: List, key_map: dict, added_by: str):
    product_raw = get_val_by_candidates(row_vals, key_map, CANDIDATES['product_name'])
    product_name = str(product_raw).strip() if product_raw not in (None, '') else None
    if not product_name:
        raise ValueError('Missing product name')

    description = get_val_by_candidates(row_vals, key_map, CANDIDATES['description']) or ''
    category = get_val_by_candidates(row_vals, key_map, CANDIDATES['category']) or ''
    unit = get_val_by_candidates(row_vals, key_map, CANDIDATES['unit']) or ''
    location = get_val_by_candidates(row_vals, key_map, CANDIDATES['location']) or ''

    raw_stock = get_val_by_candidates(row_vals, key_map, CANDIDATES['quantity_in_stock'])
    raw_use = get_val_by_candidates(row_vals, key_map, CANDIDATES['quantity_in_use'])
    raw_total = get_val_by_candidates(row_vals, key_map, CANDIDATES['quantity_total'])

    quantity_in_stock = to_int(raw_stock)
    quantity_in_use = to_int(raw_use)
    quantity_total = to_int(raw_total)

    cur = conn.cursor()

    # Try to find existing product by name + location
    cur.execute('SELECT * FROM products WHERE product_name = ? AND location = ? LIMIT 1', (product_name, location))
    existing = cur.fetchone()
    if existing:
        product_id = existing['product_id'] if 'product_id' in existing.keys() else existing[0]
        cur.execute('UPDATE products SET description = ?, category = ?, unit = ?, updated_at = CURRENT_TIMESTAMP WHERE product_id = ?', (description, category, unit, product_id))

        cur.execute('SELECT * FROM inventory_status WHERE product_id = ?', (product_id,))
        inv = cur.fetchone()
        if inv:
            old_stock = inv['quantity_in_stock'] if 'quantity_in_stock' in inv.keys() else inv[1]
            old_use = inv['quantity_in_use'] if 'quantity_in_use' in inv.keys() else inv[2]
            stock = quantity_in_stock if quantity_in_stock is not None else (old_stock or 0)
            use = quantity_in_use if quantity_in_use is not None else (old_use or 0)
            total = quantity_total if quantity_total is not None else (stock + use)
            cur.execute('UPDATE inventory_status SET quantity_in_stock = ?, quantity_in_use = ?, quantity_total = ?, last_updated = CURRENT_TIMESTAMP WHERE product_id = ?', (stock, use, total, product_id))
        else:
            stock = quantity_in_stock or 0
            use = quantity_in_use or 0
            total = quantity_total or (stock + use)
            cur.execute('INSERT INTO inventory_status (product_id, quantity_in_stock, quantity_in_use, quantity_total) VALUES (?, ?, ?, ?)', (product_id, stock, use, total))

        return ('updated', product_id)
    else:
        cur.execute('INSERT INTO products (product_name, description, category, unit, location, added_by) VALUES (?, ?, ?, ?, ?, ?)', (product_name, description, category, unit, location, added_by))
        product_id = cur.lastrowid
        stock = quantity_in_stock or 0
        use = quantity_in_use or 0
        total = quantity_total or (stock + use)
        cur.execute('INSERT INTO inventory_status (product_id, quantity_in_stock, quantity_in_use, quantity_total) VALUES (?, ?, ?, ?)', (product_id, stock, use, total))
        cur.execute('INSERT INTO transaction_history (transaction_type, product_id, quantity_change, performed_by_user_id, reference_id, reference_type, notes) VALUES (?, ?, ?, ?, ?, ?, ?)', ('PRODUCT_ADD', product_id, stock, None, product_id, 'PRODUCT', f'Imported product {product_name}'))
        return ('inserted', product_id)


def main():
    parser = argparse.ArgumentParser(description='Import inventory from Excel to SQLite DB')
    parser.add_argument('file', help='Path to Excel file (.xlsx)')
    parser.add_argument('--sheet', help='Sheet name (optional)', default=None)
    parser.add_argument('--addedBy', help='Added by text to store in products.added_by', default='importer')
    args = parser.parse_args()

    if not os.path.exists(args.file):
        raise SystemExit(f'File not found: {args.file}')

    wb = load_workbook(filename=args.file, read_only=True, data_only=True)
    sheet = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        raise SystemExit('No rows in sheet')

    key_map = build_key_map(headers)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    inserted = 0
    updated = 0
    errors = 0

    for i, r in enumerate(rows, start=2):
        row_vals = list(r)
        try:
            kind, pid = upsert_row(conn, row_vals, key_map, args.addedBy)
            conn.commit()
            if kind == 'inserted':
                inserted += 1
            else:
                updated += 1
        except Exception as e:
            errors += 1
            print(f'Row {i} skipped: {e}')

    print(f'Import complete. Inserted: {inserted}, Updated: {updated}, Errors: {errors}')
    conn.close()


if __name__ == '__main__':
    main()
